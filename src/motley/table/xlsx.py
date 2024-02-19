"""
Output tables to Excel spreadsheet
"""

# std
import contextlib as ctx
from copy import copy
from pathlib import Path
from collections import defaultdict

# third-party
import numpy as np
import more_itertools as mit
from openpyxl.cell import Cell
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

# local
from recipes import op
from recipes.string import sub
from recipes.dicts import AttrDict
from recipes.logging import LoggingMixin
from recipes.iter import cofilter, duplicates
from recipes.lists import unique, where_duplicate
from recipes.string.delimited import square_brackets
from recipes.utils import duplicate_if_scalar, ensure


# ---------------------------------------------------------------------------- #
ALIGNMENT_MAP = {'>': 'right',
                 '<': 'left',
                 '^': 'center'}
_xl_fmt_nondisplay = {'"': '', '@': ''}

# ---------------------------------------------------------------------------- #


def cell_index(i, j=''):
    n, i = divmod(i, 26)
    cell = f'{i + 65:c}{str(j)}'
    if n:
        return f'{n + 64:c}{cell}'
    return cell


def cell_range(j, r, k, s):
    return ':'.join(map(cell_index, (j, k), (r, s)))


# class hyperlink:
#     template = '=HYPERLINK("{}", "{}")'

#     def by_ext(self, path):
#         return self.template.format(path, path.suffix[1:])

#     def by_name(self, path):
#         return self.template.format(path, path.name)


# def stack_cell_attributes(cell, kws):
#     keys = ('alignment', 'font', 'border', 'fill')
#     props = op.AttrDict(*keys)(cell)
#     try:
#         # TypeError: unhashable type: 'StyleProxy'
#         op.AttrSetter(*keys)(cell, {**kws, **props})
#     except Exception as err:
#         import sys, textwrap
#         from IPython import embed
#         from better_exceptions import format_exception
#         embed(header=textwrap.dedent(
#                 f"""\
#                 Caught the following {type(err).__name__} at 'xlsx.py':25:
#                 %s
#                 Exception will be re-raised upon exiting this embedded interpreter.
#                 """) % '\n'.join(format_exception(*sys.exc_info()))
#         )
#         raise


def set_style(cell, **kws):
    # stack borders
    if (border := kws.get('border', ())):
        new = copy(cell.border)
        for atr, val in vars(border).items():
            if val and val != getattr(cell.border, atr) and val != Side():
                setattr(new, atr, val)

        kws['border'] = new

    op.AttrSetter(*kws.keys())(cell, kws.values())


def set_block_style(cells, **kws):
    itr = [cells] if isinstance(cells, Cell) else mit.flatten(cells)
    for cell in itr:
        # stack_cell_attributes(cell, kws)
        set_style(cell, **kws)
        # for key, val in kws.items():
        #     setattr(cell, key, val)


# def get_col_widths(table, requested=(), fallback=4, minimum=3):

#     # headers = table.col_headers
#     width = fallback
#     requested = mit.padded(requested, 0)
#     for i, (reqw, col) in enumerate(zip(requested, zip(*table.data))):
#         if reqw:
#             yield reqw
#             continue

#         if (fmt := table.formatters.get(i)) and isinstance(fmt, str):
#             # sub non-display characters in excel format string
#             width = max((len(sub(square_brackets.remove(f), _xl_fmt_nondisplay))
#                          for f in fmt.split(';')))
#         else:
#             # width = table.col_width[i]
#             try:
#                 width = max(map(len, col))
#             except TypeError:
#                 pass

#         header = table.col_headers[i]
#         repeats = list(table.col_headers).count(header)
#         hwidth = (len(header) * 1.2 / repeats)   # fudge factor for font
#         # logger.debug(i, header, hwidth, fwidth, width, minimum)
#         yield max(hwidth, width, minimum)


class XlsxWriter(LoggingMixin):

    # -------------------------------------------------------------------- #
    # styling defaults

    # TODO: move to config
    font = dict(name='Ubuntu Mono', size=10)
    rule = Side('thin')
    rule2 = Side('double')
    bottomrule = True

    style = {}
    style['title'] = dict(
        font=Font(**{**font,
                     'size': 14,
                     'bold': True}),
        alignment=Alignment(horizontal='center',
                            vertical='center',
                            wrap_text=True),
        border=Border(bottom=rule2)
    )

    style['headers'] = dict(
        font=Font(**{**font,
                     'size': 12,
                     'bold': True}),
        alignment=Alignment(horizontal='center',
                            vertical='center',
                            wrap_text=True),
        # border=Border(bottom=rule2)
    )

    style['units'] = dict(
        font=Font(**font),
        alignment=Alignment(horizontal='center',
                            vertical='top'),
        # border=Border(bottom=rule2)
    )
    style['data'] = dict(
        font=Font(**font),
        # alignment=Alignment(horizontal='center',
        #                     vertical='center')
    )
    style['totals'] = dict(
        font=Font(**font, bold=True),
        border=Border(bottom=rule, top=rule)
    )

    def __init__(self, table, widths=None, align=None,
                 merge_unduplicate=('headers'), header_formatter=str):

        self.logger.debug('Initializing writer.')

        if widths is None:  # is_null
            widths = {}
        if align is None:  # is_null
            align = {}

        self.table = table
        self.worksheet = None
        # worksheet.title = ""

        self.header_formatter = header_formatter

        col_widths = self.table.resolve_input(
            widths, what='width', default_factory=self.get_col_width)
        self.col_widths = np.fromiter(col_widths.values(), int)

        self.alignments = {}
        for i, al in table.resolve_input(align).items():
            if isinstance(al, str):
                self.alignments[i] = Alignment(ALIGNMENT_MAP[al])
            elif isinstance(al, dict):
                self.alignments[i] = Alignment(**al)
            else:
                raise TypeError

        # self.alignments = [) for _ in table.align]
        self.merge_unduplicate = merge_unduplicate

    def get_col_width(self, i, fallback=4, minimum=3, padding=1):
        width = fallback
        table = self.table
        if (fmt := table.formatters.get(i)) and isinstance(fmt, str):
            # sub non-display characters in excel format string
            width = max((len(sub(square_brackets.remove(f), _xl_fmt_nondisplay))
                         for f in fmt.split(';'))) + padding
        else:
            # width = table.col_width[i]
            with ctx.suppress(TypeError):
                width = max(map(len, table.data[:, i]))

        hwidth = 0
        if table.col_headers:
            header = table.col_headers[i]
            repeats = list(table.col_headers).count(header)
            hwidth = ((len(header) + 3 * padding) / repeats)
        #                        fudge factor for font
        # logger.debug(i, header, hwidth, fwidth, width, minimum)
        return max(hwidth, width, minimum)

    def write(self, path, sheet=None, formats=(), overwrite=False):
        # ('rows', 'cells')

        if path and Path(path).exists():
            workbook = load_workbook(path)
            if sheet in (None, *workbook.sheetnames) and not overwrite:
                raise FileExistsError('File at: {} exists. Overwrite is False.')

            if sheet:
                if sheet in workbook.sheetnames:
                    ws = workbook[sheet]
                    ws._current_row = 0  # !!!!
                else:
                    # new sheet
                    if workbook.sheetnames == ['Sheet']:
                        workbook.active.title = sheet
                    else:
                        ws = workbook.create_sheet(sheet)
            else:
                ws = workbook.active
        else:
            workbook = Workbook()
            ws = workbook.active
            if sheet:
                ws.title = sheet

        #
        self.worksheet = ws

        self.logger.debug('Begin writing to spreadsheet {!s}::{}.',
                          (path or ''), sheet)

        table = self.table
        nrows, ncols = table.data.shape

        # -------------------------------------------------------------------- #
        # headers
        self.make_header_block()

        # -------------------------------------------------------------------- #
        # data
        r0 = ws._current_row + 1  # first data row

        for r, row in enumerate(table.data, r0):
            self.append(row)

        r = ws._current_row

        # -------------------------------------------------------------------- #
        # Totals
        totals = [''] * ncols
        # HACK
        if isinstance(table, AttrDict):
            for i in table.totals:
                totals[i] = f'=SUM({cell_range(i, r0, i, r)})'
        elif table.totals:
            for t, i in zip(*cofilter(None, table.totals, range(ncols))):
                totals[i] = f'=SUM({cell_range(i, r0, i, r)})'

        if any(totals):
            #           data    row style
            self.append(totals, **self.style['totals'])
            r += 1

        # -------------------------------------------------------------------- #
        # Set number formats
        formats = {**table.formatters,
                   **self.table.resolve_input(formats, what='formats')}

        for idx, fmt in formats.items():
            if isinstance(fmt, str):
                # logger.debug(col, fmt)
                set_block_style(ws[cell_range(idx, r0, idx, r)],
                                number_format=fmt)

        # style for "data" cells

        set_block_style(ws[cell_range(1, r0, ncols - 1, r)], **self.style['data'])

        # set col widths
        # double_rows = defaultdict(bool)
        # z = bool(table.title) + 1
        for idx, width in enumerate(self.col_widths):
            # logger.debug(idx, table.col_headers[idx], width)
            ws.column_dimensions[cell_index(idx)].width = width
            # double_rows[z] |= ('\n' in table.col_headers[idx])
            # double_rows[z + 1] |= ('\n' in table.col_groups[idx])

        # for r, tf in double_rows.items():
        #     if tf:
        #         ws.row_dimensions[r].height = 10 * 3

        # borders
        if table.col_groups:
            for val, (*_, index) in unique(table.col_groups[0]).items():
                set_block_style(ws[cell_range(index, 1, index, r)],
                                border=Border(right=self.rule2))

            # if val in self.merge_unduplicate:
            #     self.merge_duplicate_rows( )

        if how := (ensure(set, self.merge_unduplicate) - {'headers'}):
            if 'data' in how:
                self.merge_duplicate_rows(table.data, r0, trigger=nrows)
            else:
                for h in how:
                    indices = table.resolve_columns(h, table.n_cols, 'merge region')
                    self.merge_duplicate_rows(table.data, r0, indices, nrows)

        if self.bottomrule and not table.totals:

            set_block_style(ws[cell_range(1, r + 1, ncols - 1, r + 1)],
                            border=Border(top=self.rule2))

        if path:
            workbook.save(path)
            self.logger.success('Spreadsheet saved at {!s}{}.',
                                path, (f'::{sheet}' if sheet else ''))

        return workbook

    def append(self, data, **style):
        if data is None:
            return

        sheet = self.worksheet
        sheet.append(list(data))
        r = sheet._current_row
        cells = sheet[cell_range(1, r, len(data) - 1, r)][0]

        # if self.alignments:
        for i, cell in enumerate(cells):
            if align := self.alignments.get(i):
                style.update(alignment=align)
            set_style(cell, **style)
        # else:
        #     set_block_style(cells, **style)
        return r

    def should_double_height(self, row, merged):
        unmerged = zip(set(range(len(row))) - set(sum(merged, [])))

        for i in (*merged, *unmerged):
            content, width = row[i[0]], self.col_widths[i].sum()
            if (isinstance(content, str) and
                    (('\n' in content) or len(content) > width)):

                return True

    def make_header_block(self):
        table = self.table
        ncols = table.data.shape[1]

        # title
        self.make_header_row([table.title] * ncols, **self.style['title'])
        r = self.worksheet._current_row

        # headers
        # headers = table.get_headers()
        cgroups = table.col_groups
        if isinstance(table, AttrDict):
            col_headers = list(map(self.header_formatter, table.col_headers))
            if cgroups and len(col_headers) < len(cgroups[0]):
                col_headers = ['', *col_headers]
            # col group headers
            headers = [*cgroups, col_headers]
        else:
            headers = [*cgroups, table.pre_table[0]]
            # _, headers = table.get_header_blocks(
            #     table.row_headers, table.has_row_nrs, table.col_headers)

        if table.units:
            headers.append(table.units)

        # final_style = dict(self.style['headers' if table.units else 'headers'])
        style = dict(self.style['headers'])
        for i, row in enumerate(headers):
            if i == 1:
                f = style['font'] = copy(style.pop('font'))
                f.size -= 2

            self.make_header_row(row, False, **style)

        # header borders
        q = self.worksheet._current_row
        set_block_style(

            self.worksheet[cell_range(1, q, ncols - 1, q)],
            **self.style['units' if table.units or cgroups else 'headers'],
            border=Border(bottom=self.rule2)
        )

        if 'headers' in self.merge_unduplicate:
            # for i, header in enumerate(headers, r + 1):
            self.merge_duplicate_cells(headers, r + 1)

    def make_header_row(self, data, merge_duplicate_cells=2, **style):
        if data is None:
            return

        sheet = self.worksheet
        r = sheet._current_row + 1

        # logger.debug(row)
        self.append(data)
        if merge_duplicate_cells is not False:
            self.merge_duplicate_cells(data, r, merge_duplicate_cells)

        # set style
        set_block_style(sheet[cell_range(1, r, len(data) - 1, r)],
                        **{**self.style['headers'], **style})

    def merge_duplicate_cells(self, data, row_index, trigger=2):

        data = np.atleast_2d(data)
        # logger.debug('data = {}.', data)

        nrows, _ = data.shape
        r0 = row_index
        merged = defaultdict(set)
        for r in range(nrows):
            r1 = r0 + r
            row = data[r]

            unique = {}
            duplicate = dict(duplicates(row))
            duplicate.pop('', '')
            if nrows > 1:
                unique = zip(set(range(len(row))) - set(sum(duplicate.values(), [])))

            to_merge = []
            for idx in (*duplicate.values(), *unique):
                if idx := (set(idx) - set(merged[r])):
                    to_merge.append(idx)

            # logger.debug('Row {}: to_merge {}.', r, to_merge)
            # select(to_)
            for idx in to_merge:
                j, *_, k = duplicate_if_scalar(sorted(idx), emit=False)
                # logger.debug('{}.', data[r + 1:, j:k + 1])
                extend_down = np.all(data[r + 1:, j:k + 1] == '')
                s = (nrows - r - 1) * extend_down

                # logger.debug('{}.', ( ((k - j >= trigger) | s > 0), j, k, trigger, s))
                if ((k - j + 1 >= trigger) | s > 0):
                    cells = cell_range(j, r1, k, r1 + s)
                    # logger.debug('merge: {}.', cells)
                    self.worksheet.merge_cells(cells)

                    for t in range(s + 1):
                        merged[r + t] |= idx

        # logger.debug(row, self.should_double_height(row, merged))
        # if self.should_double_height(data, merged):
        #     # style.setdefault('alignment', self.align_center_wrap)
        #     sheet.row_dimensions[sheet._current_row].height = 10 * 3

        return merged

    def merge_duplicate_rows(self, data, r0, columns=(), trigger=2, **style):
        columns = columns or range(data.shape[1])
        for i, col in enumerate(data.T):
            if i not in columns:
                continue

            col_style = dict(style)
            if (align := self.alignments.get(i)):
                col_style.update(alignment=align)

            for idx in where_duplicate(col):
                j, *_, k = idx
                # logger.debug(idx, j, k + 1)
                if idx != list(range(j, k + 1)):
                    # not sorted!
                    continue

                if len(idx) < trigger:
                    continue

                # logger.debug(i, j, k, r0)
                # logger.debug(f'merging {i:c}{j + r0}:{i:c}{k + r0}')
                cell0 = cell_index(i, j + r0)
                cell1 = cell_index(i, k + r0)
                set_style(self.worksheet[cell0], **col_style)
                self.worksheet.merge_cells(f'{cell0}:{cell1}')
